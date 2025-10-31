#!/usr/bin/env python3
"""
Script to seed financial data warehouse from Excel file (financeiro.xlsx).
Imports OMIE financial data into star schema (fact and dimension tables).

Usage:
    python scripts/seed_financial_from_excel.py
"""

import asyncio
import sys
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
import re
from typing import Dict, Optional, Any

# Add backend to path
backend_dir = Path(__file__).parent.parent
sys.path.insert(0, str(backend_dir))

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select
import os
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables from .env file
env_path = backend_dir / ".env"
load_dotenv(env_path)

from shared.models.datawarehouse import (
    DimClient, DimCategory, DimCostCenter, DimDepartment, DimDate,
    FactFinancialTransaction
)

# Excel file path
EXCEL_FILE = backend_dir / "shared" / "assets" / "financeiro.xlsx"


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def parse_date(date_val) -> Optional[date]:
    """Parse date from various formats."""
    if not date_val:
        return None

    if isinstance(date_val, datetime):
        return date_val.date()

    if isinstance(date_val, date):
        return date_val

    # Try to parse string dates
    if isinstance(date_val, str):
        # Handle DD/MM/YYYY format common in Brazilian Excel
        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                return datetime.strptime(date_val, fmt).date()
            except ValueError:
                continue

    return None


def safe_decimal(value, default=0) -> Decimal:
    """Safely convert value to Decimal."""
    if value is None or value == '':
        return Decimal(str(default))
    try:
        return Decimal(str(value))
    except:
        return Decimal(str(default))


def safe_int(value) -> Optional[int]:
    """Safely convert value to integer."""
    if value is None or value == '':
        return None
    try:
        return int(value)
    except:
        return None


def safe_str(value, max_length: Optional[int] = None) -> Optional[str]:
    """Safely convert value to string with optional max length."""
    if value is None or value == '':
        return None
    result = str(value).strip()
    if max_length and len(result) > max_length:
        result = result[:max_length]
    return result if result else None


def parse_boolean(value) -> bool:
    """Parse boolean from string."""
    if isinstance(value, bool):
        return value
    if not value:
        return False
    str_val = str(value).upper().strip()
    return str_val in ['S', 'SIM', 'YES', 'TRUE', '1', 'X']


async def get_db_url() -> str:
    """Get database URL from environment."""
    db_url = os.getenv(
        "DATABASE_URL",
        "postgresql+asyncpg://postgres:postgres@localhost:5432/cortex"
    )
    if "DATABASE_URL" in os.environ:
        print(f"📡 Using DATABASE_URL from environment")
    else:
        print(f"⚠️  DATABASE_URL not found in environment, using default")
    return db_url


# ============================================================================
# DATE DIMENSION POPULATION
# ============================================================================

async def populate_date_dimension(session: AsyncSession, start_year: int = 2020, end_year: int = 2030):
    """Populate date dimension with dates from start_year to end_year."""
    print(f"\n📅 Populating date dimension ({start_year}-{end_year})...")

    month_names_pt = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }

    day_names_pt = {
        0: 'Segunda', 1: 'Terça', 2: 'Quarta', 3: 'Quinta',
        4: 'Sexta', 5: 'Sábado', 6: 'Domingo'
    }

    current_date = date(start_year, 1, 1)
    end_date = date(end_year, 12, 31)
    dates_created = 0

    while current_date <= end_date:
        # Check if date already exists
        result = await session.execute(
            select(DimDate).where(DimDate.date_value == current_date)
        )
        if result.scalar_one_or_none():
            current_date = date(current_date.year, current_date.month, current_date.day + 1) if current_date.day < 28 else current_date.replace(day=1, month=current_date.month + 1) if current_date.month < 12 else date(current_date.year + 1, 1, 1)
            continue

        weekday = current_date.weekday()
        quarter = (current_date.month - 1) // 3 + 1

        dim_date = DimDate(
            date_value=current_date,
            year=current_date.year,
            quarter=quarter,
            month=current_date.month,
            month_name=month_names_pt[current_date.month],
            week=current_date.isocalendar()[1],
            day=current_date.day,
            day_of_week=weekday + 1,
            day_name=day_names_pt[weekday],
            is_weekend=(weekday >= 5),
            fiscal_year=current_date.year,
            fiscal_quarter=quarter,
            fiscal_period=f"{current_date.year}-{current_date.month:02d}",
            year_month=f"{current_date.year}-{current_date.month:02d}",
            year_quarter=f"{current_date.year}-Q{quarter}"
        )
        session.add(dim_date)
        dates_created += 1

        # Move to next day
        if current_date.day < 28:
            current_date = date(current_date.year, current_date.month, current_date.day + 1)
        elif current_date.month < 12:
            current_date = current_date.replace(day=1, month=current_date.month + 1)
        else:
            current_date = date(current_date.year + 1, 1, 1)

        if dates_created % 365 == 0:
            await session.flush()
            print(f"   📝 Created {dates_created} dates...")

    await session.commit()
    print(f"✅ Date dimension populated: {dates_created} dates created")


async def get_or_create_date(session: AsyncSession, date_val: Optional[date], date_cache: Dict) -> Optional[Any]:
    """Get or create date dimension entry."""
    if not date_val:
        return None

    if date_val in date_cache:
        return date_cache[date_val]

    result = await session.execute(
        select(DimDate).where(DimDate.date_value == date_val)
    )
    dim_date = result.scalar_one_or_none()

    if dim_date:
        date_cache[date_val] = dim_date.id
        return dim_date.id

    return None


# ============================================================================
# DIMENSION LOADERS
# ============================================================================

async def load_clients(session: AsyncSession, ws, col_map: Dict[str, int]) -> Dict[int, Any]:
    """Load unique clients into DimClient."""
    print(f"\n👥 Loading client dimension...")

    clients_data = {}
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row:
            continue

        omie_client_id = safe_int(row[col_map.get('nCodCliente')])
        if not omie_client_id or omie_client_id in clients_data:
            continue

        clients_data[omie_client_id] = {
            'tax_id': safe_str(row[col_map.get('cCPFCNPJCliente')], 18),
            'legal_name': safe_str(row[col_map.get('dbClientes.nome_fantasia')], 255),
            'trade_name': safe_str(row[col_map.get('nome_fantasia')], 255),
            'key_account': safe_str(row[col_map.get('Key Account')], 100),
            'vertical': safe_str(row[col_map.get('Vertical')], 100),
            'pd7_id': safe_str(row[col_map.get('PD7 ID')], 50),
            'geo': safe_str(row[col_map.get('Geo')], 100),
            'state': safe_str(row[col_map.get('Estado')], 2),
        }

    print(f"   Found {len(clients_data)} unique clients")

    client_id_map = {}
    created_count = 0

    for omie_id, client_data in clients_data.items():
        # Check if exists
        result = await session.execute(
            select(DimClient).where(DimClient.omie_client_id == omie_id)
        )
        existing = result.scalar_one_or_none()

        if existing:
            client_id_map[omie_id] = existing.id
        else:
            new_client = DimClient(
                omie_client_id=omie_id,
                **client_data
            )
            session.add(new_client)
            await session.flush()
            client_id_map[omie_id] = new_client.id
            created_count += 1

    await session.commit()
    print(f"✅ Client dimension loaded: {created_count} created, {len(clients_data) - created_count} existing")

    return client_id_map


async def load_categories(session: AsyncSession, ws, col_map: Dict[str, int]) -> Dict[str, Any]:
    """Load unique P&L categories into DimCategory."""
    print(f"\n📊 Loading category dimension...")

    categories_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        # Build category code from P&L levels
        level_value = safe_str(row[col_map.get('P&L.Level')])
        name_pt = safe_str(row[col_map.get('P&L.ptBR')])
        name_en = safe_str(row[col_map.get('P&L.enUS')])

        if not level_value or not name_pt:
            continue

        # Use name_pt as unique key
        if name_pt in categories_data:
            continue

        categories_data[name_pt] = {
            'category_code': name_pt,  # Using PT name as code
            'level': safe_int(level_value) or 1,
            'level_1': safe_str(row[col_map.get('P&L.Level_1')], 100),
            'level_2': safe_str(row[col_map.get('P&L.ptBR_2')], 100),
            'level_3': safe_str(row[col_map.get('P&L.enUS_3')], 100),
            'name_pt': name_pt,
            'name_en': name_en,
            'description': safe_str(row[col_map.get('Categorias.descricao')]),
        }

    print(f"   Found {len(categories_data)} unique categories")

    category_id_map = {}
    created_count = 0

    for category_code, category_data in categories_data.items():
        # Check if exists
        result = await session.execute(
            select(DimCategory).where(DimCategory.category_code == category_code)
        )
        existing = result.scalar_one_or_none()

        if existing:
            category_id_map[category_code] = existing.id
        else:
            new_category = DimCategory(**category_data)
            session.add(new_category)
            await session.flush()
            category_id_map[category_code] = new_category.id
            created_count += 1

    await session.commit()
    print(f"✅ Category dimension loaded: {created_count} created, {len(categories_data) - created_count} existing")

    return category_id_map


async def load_cost_centers(session: AsyncSession, ws, col_map: Dict[str, int]) -> Dict[int, Any]:
    """Load unique cost centers into DimCostCenter."""
    print(f"\n🏢 Loading cost center dimension...")

    cc_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        omie_cc_id = safe_int(row[col_map.get('nCodCC')])
        if not omie_cc_id or omie_cc_id in cc_data:
            continue

        cc_code = safe_str(row[col_map.get('cCodCateg')]) or f"CC_{omie_cc_id}"

        cc_data[omie_cc_id] = {
            'omie_cc_id': omie_cc_id,
            'cc_code': cc_code,
            'name': f"Centro de Custo {omie_cc_id}",  # Placeholder, can be enriched
            'distribution_value': safe_decimal(row[col_map.get('nDistrValor')]),
        }

    print(f"   Found {len(cc_data)} unique cost centers")

    cc_id_map = {}
    created_count = 0

    for omie_id, cc_info in cc_data.items():
        result = await session.execute(
            select(DimCostCenter).where(DimCostCenter.omie_cc_id == omie_id)
        )
        existing = result.scalar_one_or_none()

        if existing:
            cc_id_map[omie_id] = existing.id
        else:
            new_cc = DimCostCenter(**cc_info)
            session.add(new_cc)
            await session.flush()
            cc_id_map[omie_id] = new_cc.id
            created_count += 1

    await session.commit()
    print(f"✅ Cost center dimension loaded: {created_count} created, {len(cc_data) - created_count} existing")

    return cc_id_map


async def load_departments(session: AsyncSession, ws, col_map: Dict[str, int]) -> Dict[str, Any]:
    """Load unique departments into DimDepartment."""
    print(f"\n🏢 Loading department dimension...")

    dept_data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        dept_code = safe_str(row[col_map.get('cCodDepartamento')])
        dept_desc = safe_str(row[col_map.get('dbDepartamentos.descricao')])

        if not dept_code or dept_code in dept_data:
            continue

        dept_data[dept_code] = {
            'omie_dept_code': dept_code,
            'name': dept_desc or f"Departamento {dept_code}",
            'description': dept_desc,
        }

    print(f"   Found {len(dept_data)} unique departments")

    dept_id_map = {}
    created_count = 0

    for dept_code, dept_info in dept_data.items():
        result = await session.execute(
            select(DimDepartment).where(DimDepartment.omie_dept_code == dept_code)
        )
        existing = result.scalar_one_or_none()

        if existing:
            dept_id_map[dept_code] = existing.id
        else:
            new_dept = DimDepartment(**dept_info)
            session.add(new_dept)
            await session.flush()
            dept_id_map[dept_code] = new_dept.id
            created_count += 1

    await session.commit()
    print(f"✅ Department dimension loaded: {created_count} created, {len(dept_data) - created_count} existing")

    return dept_id_map


# ============================================================================
# FACT TABLE LOADER
# ============================================================================

async def load_fact_transactions(
    session: AsyncSession,
    ws,
    col_map: Dict[str, int],
    client_id_map: Dict,
    category_id_map: Dict,
    cc_id_map: Dict,
    dept_id_map: Dict,
):
    """Load financial transactions into FactFinancialTransaction."""
    print(f"\n💰 Loading fact transactions...")

    date_cache = {}
    created_count = 0
    skipped_count = 0
    error_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row:
            continue

        try:
            omie_title_id = safe_int(row[col_map.get('nCodTitulo')])
            if not omie_title_id:
                skipped_count += 1
                continue

            # Check if already exists
            result = await session.execute(
                select(FactFinancialTransaction).where(
                    FactFinancialTransaction.omie_title_id == omie_title_id
                )
            )
            if result.scalar_one_or_none():
                skipped_count += 1
                continue

            # Get dimension IDs
            omie_client_id = safe_int(row[col_map.get('nCodCliente')])
            client_id = client_id_map.get(omie_client_id)

            category_code = safe_str(row[col_map.get('P&L.ptBR')])
            category_id = category_id_map.get(category_code)

            omie_cc_id = safe_int(row[col_map.get('nCodCC')])
            cost_center_id = cc_id_map.get(omie_cc_id)

            dept_code = safe_str(row[col_map.get('cCodDepartamento')])
            department_id = dept_id_map.get(dept_code)

            # Parse dates
            issue_date = parse_date(row[col_map.get('dDtEmissao')])
            due_date = parse_date(row[col_map.get('dDtVenc')])
            payment_date = parse_date(row[col_map.get('dDtPagamento')])
            registration_date = parse_date(row[col_map.get('dDtRegistro')])

            # Get date dimension IDs
            issue_date_id = await get_or_create_date(session, issue_date, date_cache)
            due_date_id = await get_or_create_date(session, due_date, date_cache)
            payment_date_id = await get_or_create_date(session, payment_date, date_cache)
            registration_date_id = await get_or_create_date(session, registration_date, date_cache)

            # Calculate total taxes withheld
            total_taxes = sum([
                safe_decimal(row[col_map.get('nValorCOFINS')]),
                safe_decimal(row[col_map.get('nValorCSLL')]),
                safe_decimal(row[col_map.get('nValorINSS')]),
                safe_decimal(row[col_map.get('nValorIR')]),
                safe_decimal(row[col_map.get('nValorISS')]),
                safe_decimal(row[col_map.get('nValorPIS')]),
            ])

            # Create fact record
            fact = FactFinancialTransaction(
                # Dimension FKs
                client_id=client_id,
                category_id=category_id,
                cost_center_id=cost_center_id,
                department_id=department_id,
                issue_date_id=issue_date_id,
                due_date_id=due_date_id,
                payment_date_id=payment_date_id,
                registration_date_id=registration_date_id,

                # Business Keys
                omie_title_id=omie_title_id,
                omie_internal_code=safe_str(row[col_map.get('cCodIntTitulo')], 50),

                # Identifiers
                title_number=safe_str(row[col_map.get('cNumTitulo')], 100),
                parcel_number=safe_str(row[col_map.get('cNumParcela')], 20),
                fiscal_document_number=safe_str(row[col_map.get('cNumDocFiscal')], 100),

                # Classification
                transaction_type=safe_str(row[col_map.get('cTipo')], 50),
                nature=safe_str(row[col_map.get('cNatureza')], 50),
                origin=safe_str(row[col_map.get('cOrigem')], 50),
                status=safe_str(row[col_map.get('cStatus')], 50),
                is_settled=parse_boolean(row[col_map.get('cLiquidado')]),
                paid_or_received=safe_str(row[col_map.get('Pago ou Recebido')], 50),

                # Measures
                title_value=safe_decimal(row[col_map.get('nValorTitulo')]),
                net_value=safe_decimal(row[col_map.get('nValLiquido')]),
                paid_value=safe_decimal(row[col_map.get('nValPago')]),
                open_value=safe_decimal(row[col_map.get('nValAberto')]),
                discount=safe_decimal(row[col_map.get('nDesconto')]),
                interest=safe_decimal(row[col_map.get('nJuros')]),
                penalty=safe_decimal(row[col_map.get('nMulta')]),

                # Taxes
                tax_cofins=safe_decimal(row[col_map.get('nValorCOFINS')]),
                tax_csll=safe_decimal(row[col_map.get('nValorCSLL')]),
                tax_inss=safe_decimal(row[col_map.get('nValorINSS')]),
                tax_ir=safe_decimal(row[col_map.get('nValorIR')]),
                tax_iss=safe_decimal(row[col_map.get('nValorISS')]),
                tax_pis=safe_decimal(row[col_map.get('nValorPIS')]),
                total_taxes_withheld=total_taxes,

                # Tax retention flags
                has_cofins_retention=parse_boolean(row[col_map.get('cRetCOFINS')]),
                has_csll_retention=parse_boolean(row[col_map.get('cRetCSLL')]),
                has_inss_retention=parse_boolean(row[col_map.get('cRetINSS')]),
                has_ir_retention=parse_boolean(row[col_map.get('cRetIR')]),
                has_iss_retention=parse_boolean(row[col_map.get('cRetISS')]),
                has_pis_retention=parse_boolean(row[col_map.get('cRetPIS')]),

                # Linked Documents
                contract_number=safe_str(row[col_map.get('cNumCtr')], 100),
                omie_contract_id=safe_int(row[col_map.get('nCodCtr')]),
                service_order_number=safe_str(row[col_map.get('cNumOS')], 100),
                omie_service_order_id=safe_int(row[col_map.get('nCodOS')]),
                omie_invoice_id=safe_int(row[col_map.get('nCodNF')]),
                operation=safe_str(row[col_map.get('cOperacao')], 50),

                # Cost Allocation
                distribution_percentage=safe_decimal(row[col_map.get('nDistrPercentual')]),
                distribution_value=safe_decimal(row[col_map.get('nDistrValor')]),
                fixed_value=safe_decimal(row[col_map.get('nValorFixo')]),

                # Observations
                observation=safe_str(row[col_map.get('observacao')]),
                description=safe_str(row[col_map.get('descricao')]),

                # Payment Details
                payment_internal_code=safe_str(row[col_map.get('cCodIntLanc')], 50),
                payment_nature=safe_str(row[col_map.get('cNatureza.1')], 50),
                payment_observation=safe_str(row[col_map.get('cObsLanc')]),
                payment_value=safe_decimal(row[col_map.get('nValLanc')]),
                payment_discount=safe_decimal(row[col_map.get('nDesconto.1')]),
                payment_interest=safe_decimal(row[col_map.get('nJuros.1')]),
                payment_penalty=safe_decimal(row[col_map.get('nMulta.1')]),
                omie_payment_id=safe_int(row[col_map.get('nCodLanc')]),
                omie_payment_cc_id=safe_int(row[col_map.get('nIdLancCC')]),

                # Metadata
                source_row_index=row_num,
            )

            session.add(fact)
            created_count += 1

            if created_count % 500 == 0:
                await session.flush()
                print(f"   📝 Processed {created_count} transactions...")

        except Exception as e:
            print(f"   ❌ Row {row_num}: Error - {e}")
            error_count += 1
            continue

    await session.commit()

    print(f"\n✅ Fact transactions loaded:")
    print(f"   Created: {created_count}")
    print(f"   Skipped: {skipped_count}")
    print(f"   Errors: {error_count}")


# ============================================================================
# MAIN
# ============================================================================

async def seed_financial_data():
    """Main function to seed financial data warehouse."""

    # Check if Excel file exists
    if not EXCEL_FILE.exists():
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        return

    print(f"📂 Reading Excel file: {EXCEL_FILE.name}")
    print(f"   File size: {EXCEL_FILE.stat().st_size / 1024 / 1024:.1f} MB")

    # Load workbook
    try:
        print("   Loading workbook (this may take a moment for large files)...")
        wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        ws = wb.active
        print(f"✅ Loaded worksheet: {ws.title}")
        print(f"   Total rows: {ws.max_row - 1:,} (excluding header)")
        print(f"   Total columns: {ws.max_column}")
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return

    # Read headers
    headers = [cell.value for cell in ws[1]]
    col_map = {header: idx for idx, header in enumerate(headers)}

    # Create async engine
    db_url = await get_db_url()
    engine = create_async_engine(db_url, echo=False)

    async_session = sessionmaker(
        engine, class_=AsyncSession, expire_on_commit=False
    )

    async with async_session() as session:
        try:
            # Step 1: Populate date dimension
            await populate_date_dimension(session)

            # Step 2: Load dimensions
            client_id_map = await load_clients(session, ws, col_map)
            category_id_map = await load_categories(session, ws, col_map)
            cc_id_map = await load_cost_centers(session, ws, col_map)
            dept_id_map = await load_departments(session, ws, col_map)

            # Step 3: Load fact table
            await load_fact_transactions(
                session, ws, col_map,
                client_id_map, category_id_map, cc_id_map, dept_id_map
            )

            print("\n" + "=" * 60)
            print("🎉 Financial data warehouse import completed!")
            print("=" * 60)

        except Exception as e:
            await session.rollback()
            print(f"\n❌ Error importing financial data: {e}")
            import traceback
            traceback.print_exc()
            raise
        finally:
            await engine.dispose()


def main():
    """Main entry point."""
    print("=" * 60)
    print("Financial Data Warehouse Import")
    print("File: financeiro.xlsx")
    print("=" * 60)
    print()

    asyncio.run(seed_financial_data())


if __name__ == "__main__":
    main()
