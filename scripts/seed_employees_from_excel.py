#!/usr/bin/env python3
"""
Script to seed employee data from Excel file (funcionarios-pd7.xlsx).
Run this after migrations have been applied.

Usage:
    python scripts/seed_employees_from_excel.py
"""

import asyncio
import sys
from pathlib import Path
from datetime import date, datetime
from decimal import Decimal
import re

# Add backend to path
backend_dir = Path(__file__).parent.parent
sys.path.insert(0, str(backend_dir))

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select
import os
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load environment variables from .env file
env_path = backend_dir / ".env"
load_dotenv(env_path)

from shared.models.hr import Employee
from shared.models.company import Company, Department


# Excel file path
EXCEL_FILE = backend_dir / "shared" / "assets" / "funcionarios-pd7.xlsx"


def clean_cpf(cpf_str):
    """Clean CPF string to keep only digits."""
    if not cpf_str:
        return None
    # Remove any non-digit characters
    cleaned = re.sub(r'\D', '', str(cpf_str))
    # Ensure it's exactly 11 digits
    if len(cleaned) == 11:
        return cleaned
    return None


def parse_date(date_val):
    """Parse date from various formats."""
    if not date_val:
        return None

    if isinstance(date_val, datetime):
        return date_val.date()

    if isinstance(date_val, date):
        return date_val

    # Try to parse string dates
    if isinstance(date_val, str):
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(date_val, fmt).date()
            except ValueError:
                continue

    return None


def normalize_department_name(dept_name):
    """Normalize department/allocation name."""
    if not dept_name:
        return None

    # Clean up the name
    name = str(dept_name).strip()

    # Remove codes like "0010 - " from beginning
    name = re.sub(r'^\d+\s*-\s*', '', name)

    # Limit length
    if len(name) > 100:
        name = name[:100]

    return name if name else None


def map_contract_type(regime):
    """Map regime de contratação to our contract types."""
    if not regime:
        return "CLT"

    regime = str(regime).upper().strip()

    if "CLT" in regime:
        return "CLT"
    elif "PJ" in regime or "PRESTADOR" in regime:
        return "PJ"
    elif "ESTAG" in regime or "INTERN" in regime:
        return "Estagio"

    return "CLT"  # Default


def determine_work_mode(alocacao):
    """Determine work mode based on allocation."""
    if not alocacao:
        return "Presencial"

    alocacao = str(alocacao).upper()

    if "REMOTO" in alocacao or "HOME" in alocacao:
        return "Remoto"
    elif "HIBRID" in alocacao:
        return "Hibrido"

    return "Presencial"  # Default


async def get_db_url() -> str:
    """Get database URL from environment."""
    db_url = os.getenv(
        "DATABASE_URL",
        "postgresql+asyncpg://postgres:postgres@localhost:5432/cortex"
    )

    if "DATABASE_URL" in os.environ:
        print(f"📡 Using DATABASE_URL from environment")
    else:
        print(f"⚠️  DATABASE_URL not found in environment, using default")

    return db_url


async def seed_employees_from_excel():
    """Seed employee data from Excel file."""

    # Check if Excel file exists
    if not EXCEL_FILE.exists():
        print(f"❌ Excel file not found: {EXCEL_FILE}")
        print(f"   Please ensure the file exists at the specified location.")
        return

    print(f"📂 Reading Excel file: {EXCEL_FILE.name}")

    # Load workbook
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        print(f"✅ Loaded worksheet: {ws.title}")
        print(f"   Total rows: {ws.max_row - 1} (excluding header)")
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        return

    # Read headers
    headers = [cell.value for cell in ws[1]]
    print(f"✅ Found {len(headers)} columns")

    # Create column index mapping
    col_map = {header: idx for idx, header in enumerate(headers)}

    # Create async engine
    db_url = await get_db_url()
    engine = create_async_engine(db_url, echo=False)  # Set echo=False for cleaner output

    # Create async session factory
    async_session = sessionmaker(
        engine, class_=AsyncSession, expire_on_commit=False
    )

    async with async_session() as session:
        try:
            # Get the first company
            result = await session.execute(
                select(Company).limit(1)
            )
            company = result.scalar_one_or_none()

            if not company:
                print("❌ No company found in database!")
                print("   Please run migrations and ensure at least one company exists.")
                return

            print(f"✅ Found company: {company.company_name} (ID: {company.id})")

            # Track departments we'll create
            departments_to_create = {}
            department_cache = {}

            # First pass: collect unique departments from Excel
            print(f"\n📊 Analyzing departments from Excel...")
            unique_depts = set()

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[col_map.get('CPF')]:  # Skip empty rows
                    continue

                alocacao = row[col_map.get('Alocação')]
                dept_name = normalize_department_name(alocacao)

                if dept_name:
                    unique_depts.add(dept_name)

            print(f"✅ Found {len(unique_depts)} unique departments")

            # Create departments
            print(f"\n🏢 Creating departments...")
            created_dept_count = 0

            for dept_name in sorted(unique_depts):
                # Check if department already exists
                result = await session.execute(
                    select(Department).where(
                        Department.company_id == company.id,
                        Department.name == dept_name
                    )
                )
                existing_dept = result.scalar_one_or_none()

                if existing_dept:
                    department_cache[dept_name] = existing_dept.id
                    print(f"   ⚠️  Department '{dept_name}' already exists")
                else:
                    # Create new department
                    new_dept = Department(
                        company_id=company.id,
                        name=dept_name,
                        metadata={"created_by": "seed_script", "source": "excel_import"}
                    )
                    session.add(new_dept)
                    await session.flush()  # Get the ID
                    department_cache[dept_name] = new_dept.id
                    created_dept_count += 1
                    print(f"   ✅ Created department: {dept_name}")

            await session.commit()
            print(f"✅ Departments ready: {len(department_cache)} total ({created_dept_count} created, {len(unique_depts) - created_dept_count} existing)")

            # Second pass: create employees
            print(f"\n👥 Creating employees...")
            created_count = 0
            skipped_count = 0
            error_count = 0
            active_only = True  # Set to False to import all employees

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                if not row:
                    continue

                try:
                    # Extract data from row
                    nome = row[col_map.get('Nome')]
                    cpf_raw = row[col_map.get('CPF')]
                    status = row[col_map.get('Status:')]

                    # Skip if no CPF or name
                    if not cpf_raw or not nome:
                        continue

                    # Clean CPF
                    cpf = clean_cpf(cpf_raw)
                    if not cpf:
                        print(f"   ⚠️  Row {row_num}: Invalid CPF '{cpf_raw}' - skipping")
                        skipped_count += 1
                        continue

                    # Filter by status if needed
                    if active_only and status != "Ativo":
                        skipped_count += 1
                        continue

                    # Check if employee already exists
                    result = await session.execute(
                        select(Employee).where(Employee.tax_id == cpf)
                    )
                    if result.scalar_one_or_none():
                        skipped_count += 1
                        continue

                    # Extract other fields
                    funcao = row[col_map.get('Função')] or row[col_map.get('Cargo')] or "Não especificado"
                    regime = row[col_map.get('Regime de Contratação:')]
                    alocacao = row[col_map.get('Alocação')]
                    data_admissao = parse_date(row[col_map.get('Data de Admissão:')])
                    email_pessoal = row[col_map.get('Email  Pessoal:')]
                    email_pd7 = row[col_map.get('Email PD7:')]
                    telefone = row[col_map.get('Telefone principal')]
                    telefone_alt = row[col_map.get('Telefone Alternativo')]
                    data_nascimento = parse_date(row[col_map.get('Data de nascimento')])
                    rg = row[col_map.get('RG')]
                    tipo_sanguineo = row[col_map.get('Tipo Sanguíneo')]
                    escolaridade = row[col_map.get('Grau de escolaridade')]
                    endereco = row[col_map.get('Endereço')]

                    # Get department ID
                    dept_name = normalize_department_name(alocacao)
                    department_id = department_cache.get(dept_name)

                    # Prepare employee data
                    employee_data = {
                        "company_id": company.id,
                        "department_id": department_id,
                        "name": str(nome)[:255] if nome else None,
                        "tax_id": cpf,
                        "position": str(funcao)[:100] if funcao else "Não especificado",
                        "hire_date": data_admissao or date.today(),
                        "birth_date": data_nascimento,
                        "contract_type": map_contract_type(regime),
                        "work_mode": determine_work_mode(alocacao),
                        "status": "active" if status == "Ativo" else "inactive",
                        "personal_data": {
                            "full_name": str(nome),
                            "email_pessoal": str(email_pessoal) if email_pessoal else None,
                            "email_empresa": str(email_pd7) if email_pd7 else None,
                            "phone": str(telefone) if telefone else None,
                            "phone_alt": str(telefone_alt) if telefone_alt else None,
                            "tipo_sanguineo": str(tipo_sanguineo) if tipo_sanguineo else None,
                            "escolaridade": str(escolaridade) if escolaridade else None,
                            "address": str(endereco) if endereco else None
                        },
                        "documents": {
                            "cpf": cpf,
                            "rg": str(rg) if rg else None
                        },
                        "vacation": {
                            "available_days": 30,
                            "history": []
                        }
                    }

                    # Create employee
                    employee = Employee(**employee_data)
                    session.add(employee)
                    created_count += 1

                    if created_count % 100 == 0:
                        await session.flush()
                        print(f"   📝 Processed {created_count} employees...")

                except Exception as e:
                    print(f"   ❌ Row {row_num}: Error - {e}")
                    error_count += 1
                    continue

            # Final commit
            await session.commit()

            print("\n" + "=" * 60)
            print(f"🎉 Employee import completed!")
            print(f"   Created: {created_count}")
            print(f"   Skipped: {skipped_count}")
            print(f"   Errors: {error_count}")
            print(f"   Departments: {len(department_cache)}")
            print("=" * 60)

        except Exception as e:
            await session.rollback()
            print(f"\n❌ Error importing employees: {e}")
            print("\n💡 Troubleshooting tips:")
            print("   1. Ensure PostgreSQL is running: docker-compose up -d postgres")
            print("   2. Verify DATABASE_URL in .env file")
            print("   3. Check database credentials are correct")
            print("   4. Run migrations first: alembic upgrade head")
            raise
        finally:
            await engine.dispose()


def main():
    """Main entry point."""
    print("=" * 60)
    print("Employee Import from Excel")
    print(f"File: funcionarios-pd7.xlsx")
    print("=" * 60)
    print()

    asyncio.run(seed_employees_from_excel())


if __name__ == "__main__":
    main()
